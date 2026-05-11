"""Inspect the three Excel source files: list sheets, columns, dtypes, sample rows."""
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

DESKTOP = Path("/mnt/c/Users/wesley/Desktop")
FILES = {
    "BASE":         DESKTOP / "00 - Base.xlsx",
    "CLASSIF":      DESKTOP / "01 - Classificação.xlsx",
    "RATEADOS":     DESKTOP / "02 - Valores Rateados.xlsx",
}

for label, path in FILES.items():
    print("=" * 90)
    print(f"### {label}  →  {path.name}   ({path.stat().st_size/1024/1024:.2f} MB)")
    print("=" * 90)
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\n  -- Aba: '{sheet}'  (dim={ws.calculate_dimension()}, max_row≈{ws.max_row}, max_col={ws.max_column})")
    wb.close()

    # Read each sheet (head only) for column listing
    xl = pd.ExcelFile(path, engine="openpyxl")
    for sheet in xl.sheet_names:
        try:
            df = pd.read_excel(xl, sheet_name=sheet, nrows=5)
            print(f"\n   [{sheet}] colunas ({len(df.columns)}):")
            for c in df.columns:
                print(f"       - {c!r}  dtype={df[c].dtype}")
            print(f"   primeiras linhas:")
            print(df.head(3).to_string(index=False, max_colwidth=40))
        except Exception as e:
            print(f"   [{sheet}] erro: {e}")
    print()
