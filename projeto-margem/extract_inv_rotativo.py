#!/usr/bin/env python3
"""Lê 'Inventário Rotativo.xlsx' e gera data/inv_rotativo.json.

Sheets esperadas:
- "Inventário": ANO, MES, NROEMPRESA, COMPRADOR, CAMINHOCOMPLETO, SEQPRODUTO,
                PRODUTO, QTD_DIFERENCA, VLR_DIFERENCA
- "Vendas":     ANO, MES, NROEMPRESA, APELIDO, CAMINHOCOMPLETO, SEQPRODUTO,
                DESCCOMPLETA, VENDA, MARGEM, VERBA

O JSON agrega por NROEMPRESA pegando o período (ano, mês) mais recente
presente na sheet Inventário.

Saída (data/inv_rotativo.json):
  {
    "gerado_em": "...",
    "periodo": {"ano": 2026, "mes": "04"},
    "lojas": [
      {"nroempresa": 5, "valor": -5756.04, "qtd": -1638.4,
       "venda": 499328.21, "pct": -0.0115}
    ],
    "total": {"valor": ..., "qtd": ..., "venda": ..., "pct": ...},
    "itens": [
      {"nroempresa": 5, "comprador": "01-MAURIC(SEC)",
       "secao": "MERCEARIA \\\\ MERCEARIA \\\\ BISCOITOS",
       "seqproduto": 6718, "produto": "TORRADA 120G ...",
       "qtd": 13, "valor": 52.52}
    ]
  }
"""
import json, sys, os
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl

XLSX = os.environ.get('INV_ROTATIVO_XLSX', 'data/Inventario_Rotativo.xlsx')
OUT  = Path('data/inv_rotativo.json')

def to_int(v):
    try: return int(float(v))
    except: return None
def to_float(v):
    try: return float(v)
    except: return None

def main():
    if not Path(XLSX).exists():
        sys.stderr.write(f'ERRO: planilha não encontrada: {XLSX}\n')
        return 1
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    if 'Inventário' not in wb.sheetnames or 'Vendas' not in wb.sheetnames:
        sys.stderr.write(f'ERRO: sheets esperadas "Inventário" e "Vendas" não encontradas. Sheets: {wb.sheetnames}\n')
        return 2

    # ===== Inventário =====
    ws = wb['Inventário']
    header = None
    inv_rows = []
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = row; continue
        if row is None or all(v is None for v in row): continue
        d = dict(zip(header, row))
        ano = to_int(d.get('ANO')); mes = d.get('MES'); nro = to_int(d.get('NROEMPRESA'))
        if not (ano and mes and nro): continue
        inv_rows.append({
            'ano': ano, 'mes': str(mes).zfill(2),
            'nroempresa': nro,
            'comprador': d.get('COMPRADOR') or '',
            'secao':     d.get('CAMINHOCOMPLETO') or '',
            'seqproduto': to_int(d.get('SEQPRODUTO')),
            'produto':   d.get('PRODUTO') or '',
            'qtd':       to_float(d.get('QTD_DIFERENCA')) or 0,
            'valor':     to_float(d.get('VLR_DIFERENCA')) or 0,
        })
    print(f'Inventário: {len(inv_rows)} linhas', file=sys.stderr)

    # ===== Vendas (pra calcular % por loja) =====
    ws = wb['Vendas']
    header = None
    venda = defaultdict(float)  # (ano,mes,nroempresa) -> sum(VENDA)
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = row; continue
        if row is None or all(v is None for v in row): continue
        d = dict(zip(header, row))
        ano = to_int(d.get('ANO')); mes = d.get('MES'); nro = to_int(d.get('NROEMPRESA'))
        if not (ano and mes and nro): continue
        venda[(ano, str(mes).zfill(2), nro)] += to_float(d.get('VENDA')) or 0
    print(f'Vendas: {len(venda)} agregados (ano,mes,empresa)', file=sys.stderr)

    wb.close()

    # ===== Período mais recente =====
    periodos = sorted({(r['ano'], r['mes']) for r in inv_rows}, reverse=True)
    if not periodos:
        sys.stderr.write('Sem períodos em Inventário\n'); return 3
    ano_atual, mes_atual = periodos[0]
    print(f'Período mais recente: {ano_atual}/{mes_atual}', file=sys.stderr)

    inv_periodo = [r for r in inv_rows if r['ano'] == ano_atual and r['mes'] == mes_atual]

    # ===== Agrega por loja =====
    por_loja = defaultdict(lambda: {'valor': 0.0, 'qtd': 0.0})
    for r in inv_periodo:
        por_loja[r['nroempresa']]['valor'] += r['valor']
        por_loja[r['nroempresa']]['qtd']   += r['qtd']

    lojas = []
    for nro, agg in sorted(por_loja.items()):
        v = venda.get((ano_atual, mes_atual, nro), 0.0)
        lojas.append({
            'nroempresa': nro,
            'valor': round(agg['valor'], 2),
            'qtd':   round(agg['qtd'], 3),
            'venda': round(v, 2),
            'pct':   (agg['valor'] / v) if v else None,
        })

    tot_v = sum(l['valor'] for l in lojas)
    tot_q = sum(l['qtd']   for l in lojas)
    tot_x = sum(l['venda'] for l in lojas)
    total = {
        'valor': round(tot_v, 2),
        'qtd':   round(tot_q, 3),
        'venda': round(tot_x, 2),
        'pct':   (tot_v / tot_x) if tot_x else None,
    }

    itens = [{
        'nroempresa': r['nroempresa'],
        'comprador':  r['comprador'],
        'secao':      r['secao'],
        'seqproduto': r['seqproduto'],
        'produto':    r['produto'],
        'qtd':        round(r['qtd'], 3),
        'valor':      round(r['valor'], 2),
    } for r in inv_periodo]

    out = {
        'gerado_em': datetime.now().isoformat(timespec='seconds'),
        'periodo':   {'ano': ano_atual, 'mes': mes_atual},
        'lojas':     lojas,
        'total':     total,
        'itens':     itens,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, ensure_ascii=False, default=str), encoding='utf-8')
    print(f'✓ salvo em {OUT} ({OUT.stat().st_size // 1024} KB)', file=sys.stderr)
    print(f'Lojas: {len(lojas)} · Itens: {len(itens)} · Total valor: {total["valor"]:.2f}', file=sys.stderr)
    return 0

if __name__ == '__main__':
    sys.exit(main())
